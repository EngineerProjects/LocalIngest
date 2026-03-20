"""
Génération du rapport Excel — Étape 4.

- aggregate_anomalies_by_site : enrichit le DataFrame avec les flags d'anomalies
- compute_global_stats : calcule les statistiques globales
- generate_excel_report : crée le fichier Excel avec 3 onglets
"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.config import AnomalyCode, Config, Severity
from src.models import AnomalyCollector
from src.utils import format_number


# =============================================================================
# AGRÉGATION DES ANOMALIES PAR SITE
# =============================================================================

def aggregate_anomalies_by_site(
    df: pd.DataFrame,
    collector: AnomalyCollector,
) -> pd.DataFrame:
    """
    Enrichit le DataFrame source avec les informations d'anomalies agrégées par site.
    Ajoute les colonnes : HAS_ANOMALY, ANOMALY_COUNT, ANOMALY_CODES, WORST_SEVERITY.
    """
    print("\n" + "=" * 60)
    print("   AGRÉGATION DES RÉSULTATS")
    print("=" * 60)

    df_anomalies = collector.to_dataframe()

    if df_anomalies.empty:
        df["HAS_ANOMALY"] = False
        df["ANOMALY_COUNT"] = 0
        df["ANOMALY_CODES"] = ""
        df["WORST_SEVERITY"] = ""
        return df

    # Agréger par site
    severity_order = {Severity.GRAVE: 3, Severity.LEGERE: 2, Severity.INFO: 1}

    def get_worst_severity(severities):
        return max(severities, key=lambda x: severity_order.get(x, 0))

    agg = df_anomalies.groupby("ID_SITE").agg(
        ANOMALY_COUNT=("CODE_ANOMALIE", "count"),
        ANOMALY_CODES=("CODE_ANOMALIE", lambda x: ", ".join(sorted(set(x)))),
        WORST_SEVERITY=("GRAVITE", get_worst_severity),
    ).reset_index()

    # Joindre au DataFrame source
    df = df.merge(agg, left_on=Config.COL_SITE_ID, right_on="ID_SITE", how="left")

    # Remplir les valeurs manquantes
    df["HAS_ANOMALY"] = df["ANOMALY_COUNT"].notna()
    df["ANOMALY_COUNT"] = df["ANOMALY_COUNT"].fillna(0).astype(int)
    df["ANOMALY_CODES"] = df["ANOMALY_CODES"].fillna("")
    df["WORST_SEVERITY"] = df["WORST_SEVERITY"].fillna("")

    # Supprimer la colonne ID_SITE dupliquée si présente
    if "ID_SITE_y" in df.columns:
        df = df.drop(columns=["ID_SITE_y"])
    if "ID_SITE_x" in df.columns:
        df = df.rename(columns={"ID_SITE_x": "ID_SITE"})

    # Statistiques
    sites_with_anomalies = df["HAS_ANOMALY"].sum()
    sites_without_anomalies = len(df) - sites_with_anomalies

    print(f"\n📊 Résultats :")
    print(f"   Sites avec anomalie(s)  : {format_number(int(sites_with_anomalies))}")
    print(f"   Sites sans anomalie     : {format_number(int(sites_without_anomalies))}")
    print(f"   Total anomalies         : {format_number(collector.count())}")

    return df


# =============================================================================
# STATISTIQUES GLOBALES
# =============================================================================

def compute_global_stats(
    df: pd.DataFrame,
    collector: AnomalyCollector,
) -> dict:
    """Calcule les statistiques globales pour le rapport."""
    return {
        "date_execution": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_lignes": len(df),
        "lignes_avec_anomalies": int(df["HAS_ANOMALY"].sum()),
        "lignes_sans_anomalies": int((~df["HAS_ANOMALY"]).sum()),
        "total_anomalies": collector.count(),
        "anomalies_par_gravite": collector.count_by_severity(),
        "anomalies_par_code": collector.count_by_code(),
    }


# =============================================================================
# GÉNÉRATION DU RAPPORT EXCEL
# =============================================================================

def generate_excel_report(
    df_enriched: pd.DataFrame,
    collector: AnomalyCollector,
    stats: dict,
    output_path: Path = None,
) -> Path:
    """
    Génère le rapport Excel complet avec 3 onglets :
    - Résumé : statistiques globales + répartition par code
    - Détail anomalies : une ligne par anomalie
    - Données enrichies : fichier source avec flags
    """
    print("\n" + "=" * 60)
    print("   GÉNÉRATION DU RAPPORT EXCEL")
    print("=" * 60)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Config.OUTPUT_DIR / f"rapport_anomalies_{timestamp}.xlsx"

    # Créer le répertoire de sortie si nécessaire
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # ----- Onglet 1 : Résumé -----
    ws_resume = wb.active
    ws_resume.title = "Résumé"

    # Section A : Statistiques globales
    ws_resume["A1"] = "STATISTIQUES GLOBALES"
    ws_resume["A1"].font = Font(bold=True, size=14)

    stats_data = [
        ("Date d'exécution", stats["date_execution"]),
        ("Total lignes analysées", format_number(stats["total_lignes"])),
        ("Lignes sans anomalie", format_number(stats["lignes_sans_anomalies"])),
        ("Lignes avec anomalie(s)", format_number(stats["lignes_avec_anomalies"])),
        ("Total anomalies détectées", format_number(stats["total_anomalies"])),
        ("Anomalies GRAVES", format_number(stats["anomalies_par_gravite"].get(Severity.GRAVE, 0))),
        ("Anomalies LÉGÈRES", format_number(stats["anomalies_par_gravite"].get(Severity.LEGERE, 0))),
        ("Anomalies INFO", format_number(stats["anomalies_par_gravite"].get(Severity.INFO, 0))),
    ]

    for i, (label, value) in enumerate(stats_data, start=3):
        ws_resume[f"A{i}"] = label
        ws_resume[f"B{i}"] = value

    # Section B : Répartition par code
    start_row = len(stats_data) + 5
    ws_resume[f"A{start_row}"] = "RÉPARTITION PAR CODE ANOMALIE"
    ws_resume[f"A{start_row}"].font = Font(bold=True, size=14)

    ws_resume[f"A{start_row + 2}"] = "Code"
    ws_resume[f"B{start_row + 2}"] = "Libellé"
    ws_resume[f"C{start_row + 2}"] = "Occurrences"
    ws_resume[f"D{start_row + 2}"] = "Gravité"

    for col_letter in ["A", "B", "C", "D"]:
        ws_resume[f"{col_letter}{start_row + 2}"].font = header_font
        ws_resume[f"{col_letter}{start_row + 2}"].fill = header_fill

    for i, (code, count) in enumerate(
        stats["anomalies_par_code"].items(), start=start_row + 3
    ):
        ws_resume[f"A{i}"] = code
        ws_resume[f"B{i}"] = AnomalyCode.get_label(code)
        ws_resume[f"C{i}"] = count
        ws_resume[f"D{i}"] = AnomalyCode.get_severity(code)

    # Largeurs de colonnes
    ws_resume.column_dimensions["A"].width = 30
    ws_resume.column_dimensions["B"].width = 50
    ws_resume.column_dimensions["C"].width = 15
    ws_resume.column_dimensions["D"].width = 12

    # ----- Onglet 2 : Détail des anomalies -----
    ws_detail = wb.create_sheet("Détail anomalies")

    df_anomalies = collector.to_dataframe()

    if not df_anomalies.empty:
        for col_idx, col_name in enumerate(df_anomalies.columns, start=1):
            cell = ws_detail.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row in enumerate(df_anomalies.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws_detail.cell(row=row_idx, column=col_idx, value=value)

    # ----- Onglet 3 : Données enrichies -----
    ws_data = wb.create_sheet("Données enrichies")

    cols_to_export = [
        Config.COL_SITE_ID,
        Config.COL_CONTRACT_ID,
        Config.COL_COUNTRY,
        Config.COL_POSTAL_CODE,
        Config.COL_CITY,
        Config.COL_LONGITUDE,
        Config.COL_LATITUDE,
        "HAS_ANOMALY",
        "ANOMALY_COUNT",
        "ANOMALY_CODES",
        "WORST_SEVERITY",
    ]

    cols_to_export = [c for c in cols_to_export if c in df_enriched.columns]
    df_export = df_enriched[cols_to_export]

    # En-têtes
    for col_idx, col_name in enumerate(df_export.columns, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Données (limiter à 100K lignes pour Excel)
    max_rows = min(len(df_export), 100_000)
    if len(df_export) > max_rows:
        print(f"   ⚠️ Limitation à {format_number(max_rows)} lignes pour l'onglet données")

    for row_idx, row in enumerate(
        df_export.head(max_rows).itertuples(index=False), start=2
    ):
        for col_idx, value in enumerate(row, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    # Sauvegarder
    wb.save(output_path)

    print(f"\n✅ Rapport généré : {output_path}")
    print(f"   📊 Onglet 'Résumé' : statistiques globales")
    print(f"   📋 Onglet 'Détail anomalies' : {format_number(len(df_anomalies))} lignes")
    print(f"   📁 Onglet 'Données enrichies' : {format_number(max_rows)} lignes")

    return output_path
