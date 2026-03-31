from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.config import INPUT_FILE, OUTPUT_DIR, IssueCode, ValidationMode
from src.utils import format_number


# ==============================================================================
# PALETTE DE STYLES
# ==============================================================================

_COLOR_GREEN_DARK  = "2F5D50"   # en-têtes tableau
_COLOR_GREEN_LIGHT = "EAF2EE"   # fond titres de section
_COLOR_ROW_ALT     = "F5F9F7"   # rangées alternées tableau

def _font(size=11, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _align(h="left", v="top", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_thin():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def _border_header():
    side = Side(style="medium", color=_COLOR_GREEN_DARK)
    return Border(left=side, right=side, top=side, bottom=side)

# Styles pré-définis
_STYLE_HEADER = {
    "font":   _font(size=11, bold=True, color="FFFFFF"),
    "fill":   _fill(_COLOR_GREEN_DARK),
    "align":  _align(h="center", v="center"),
    "border": _border_header(),
}
_STYLE_SECTION_TITLE = {
    "font":   _font(size=13, bold=True, color=_COLOR_GREEN_DARK),
    "fill":   _fill(_COLOR_GREEN_LIGHT),
    "align":  _align(v="center"),
}
_STYLE_KV_LABEL = {
    "font":  _font(size=11, bold=True),
    "align": _align(v="center"),
}
_STYLE_KV_LABEL_INDENT = {
    "font":  _font(size=11, color="555555"),
    "align": _align(v="center"),
}
_STYLE_KV_VALUE = {
    "font":  _font(size=11),
    "align": _align(h="right", v="center"),
}
_STYLE_BODY = {
    "font":   _font(size=11),
    "align":  _align(v="top", wrap=True),
    "border": _border_thin(),
}
_STYLE_BODY_ALT = {
    "font":   _font(size=11),
    "fill":   _fill(_COLOR_ROW_ALT),
    "align":  _align(v="top", wrap=True),
    "border": _border_thin(),
}
_STYLE_PCT = {
    "font":   _font(size=11),
    "align":  _align(h="center", v="top"),
    "border": _border_thin(),
}
_STYLE_PCT_ALT = {
    "font":   _font(size=11),
    "fill":   _fill(_COLOR_ROW_ALT),
    "align":  _align(h="center", v="top"),
    "border": _border_thin(),
}


def _apply(cell, style: dict):
    """Applique un dictionnaire de styles à une cellule."""
    if "font"   in style: cell.font      = style["font"]
    if "fill"   in style: cell.fill      = style["fill"]
    if "align"  in style: cell.alignment = style["align"]
    if "border" in style: cell.border    = style["border"]


# ==============================================================================
# STATISTIQUES GLOBALES
# ==============================================================================

def compute_global_stats(
    df: pd.DataFrame,
    nb_total_fichier: int = 0,
    nb_inactifs: int = 0,
) -> dict:
    """
    Calcule les statistiques globales du pipeline.

    Args:
        df               : DataFrame enrichi (sites actifs analysés)
        nb_total_fichier : nombre de lignes dans le fichier Proxima source
        nb_inactifs      : nombre de sites exclus par le filtre contrats actifs
    """
    issue_counts = Counter()
    if "_ISSUE_CODES" in df.columns:
        for codes in df["_ISSUE_CODES"].fillna(""):
            for code in [c.strip() for c in str(codes).split(",") if c.strip()]:
                issue_counts[code] += 1

    return {
        "date_execution":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "nb_total_fichier": nb_total_fichier,
        "nb_inactifs":      nb_inactifs,
        "nb_analyses":      len(df),
        "nb_sites_ok":      int((df["_ISSUE_COUNT"] == 0).sum()),
        "nb_gps_ok":        int((df["_CONTROL_MODE"] == ValidationMode.GPS).sum()),
        "nb_adresse_ok":    int((df["_CONTROL_MODE"] == ValidationMode.ADDRESS).sum()),
        "nb_avec_anomalies": int((df["_ISSUE_COUNT"] > 0).sum()),
        "issue_counts":     dict(issue_counts.most_common()),
    }


# ==============================================================================
# GÉNÉRATION DU RAPPORT
# ==============================================================================

def generate_excel_report(
    df: pd.DataFrame,
    stats: dict,
    output_path: Path = None,
) -> Path:
    """Génère le rapport Excel (2 onglets : Résumé + Détail anomalies)."""
    print("\n" + "=" * 60)
    print("   GÉNÉRATION DU RAPPORT EXCEL")
    print("=" * 60)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / f"rapport_anomalies_{timestamp}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    _write_summary_sheet(wb.active, stats)

    df_anomalies = _build_anomalies_detail(df)
    _write_anomalies_sheet(wb.create_sheet("Détail anomalies"), df_anomalies)

    wb.save(output_path)

    print(f"\nRapport généré : {output_path}")
    print(f"   Sites avec anomalies : {format_number(len(df_anomalies))}")

    return output_path


# ==============================================================================
# CONSTRUCTION DES DONNÉES
# ==============================================================================

def _build_anomalies_detail(df: pd.DataFrame) -> pd.DataFrame:
    """
    Construit la vue détail anomalies : 1 ligne par site avec anomalie.

    Colonnes : ID_SITE | NU_CNT | LIBELLE | DESCRIPTION
    """
    rows = []

    for _, row in df.iterrows():
        issue_summary = row.get("_ISSUE_SUMMARY", "")
        issue_details = row.get("_ISSUE_DETAILS", [])

        if not issue_summary:
            continue

        first_detail = issue_details[0] if issue_details else {}

        rows.append({
            "ID_SITE":     row.get("ID_SITE"),
            "NU_CNT":      row.get("NU_CNT"),
            "LIBELLE":     issue_summary,
            "DESCRIPTION": first_detail.get("DESCRIPTION", ""),
        })

    if not rows:
        return pd.DataFrame(columns=["ID_SITE", "NU_CNT", "LIBELLE", "DESCRIPTION"])

    return pd.DataFrame(rows)[["ID_SITE", "NU_CNT", "LIBELLE", "DESCRIPTION"]]


# ==============================================================================
# ONGLET RÉSUMÉ
# ==============================================================================

def _write_summary_sheet(ws, stats: dict) -> None:
    """
    Écrit l'onglet Résumé avec mise en forme soignée.

    Section A — Vue globale
    Section B — Répartition par anomalie
    """
    ws.title = "Résumé"
    ws.sheet_view.showGridLines = False

    row = 1

    def blank(n=1):
        nonlocal row
        row += n

    def write_section_title(text: str):
        nonlocal row
        cell = ws.cell(row=row, column=1, value=text)
        _apply(cell, _STYLE_SECTION_TITLE)
        # Fusionner les deux colonnes pour le titre
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 26
        row += 1

    def write_kv(label: str, value, indent: bool = False):
        nonlocal row
        prefix = "    " if indent else ""
        label_cell = ws.cell(row=row, column=1, value=f"{prefix}{label}")
        value_cell = ws.cell(row=row, column=2, value=value)
        _apply(label_cell, _STYLE_KV_LABEL_INDENT if indent else _STYLE_KV_LABEL)
        _apply(value_cell, _STYLE_KV_VALUE)
        ws.row_dimensions[row].height = 20
        row += 1

    def write_table_header(values):
        nonlocal row
        for col_idx, text in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=text)
            _apply(cell, _STYLE_HEADER)
        ws.row_dimensions[row].height = 22
        row += 1

    # ------------------------------------------------------------------
    # SECTION A
    # ------------------------------------------------------------------
    write_section_title("SECTION A — Vue globale")
    blank()

    write_kv("Date d'exécution",            stats["date_execution"])
    write_kv("Fichier source",              INPUT_FILE.name)
    blank()

    write_kv("Sites dans le fichier source",    format_number(stats["nb_total_fichier"]))
    write_kv("Sites exclus (contrat inactif)",   format_number(stats["nb_inactifs"]))
    blank()

    write_kv("Sites actifs analysés",           format_number(stats["nb_analyses"]))
    blank()

    write_kv("Sites sans anomalie",             format_number(stats["nb_sites_ok"]))
    write_kv("- validés par GPS",               format_number(stats["nb_gps_ok"]),     indent=True)
    write_kv("- validés par adresse",           format_number(stats["nb_adresse_ok"]), indent=True)
    blank()

    write_kv("Sites avec anomalie(s)",          format_number(stats["nb_avec_anomalies"]))

    # ------------------------------------------------------------------
    # SECTION B
    # ------------------------------------------------------------------
    blank(2)
    write_section_title("SECTION B — Répartition par anomalie")
    blank()

    write_table_header(["Libellé de l'anomalie", "Occurrences", "%"])

    nb_analyses = stats["nb_analyses"]
    for i, (code, count) in enumerate(stats["issue_counts"].items()):
        pct = (count / nb_analyses * 100) if nb_analyses > 0 else 0
        style_body = _STYLE_BODY_ALT if i % 2 == 1 else _STYLE_BODY
        style_pct  = _STYLE_PCT_ALT  if i % 2 == 1 else _STYLE_PCT

        cell_lib  = ws.cell(row=row, column=1, value=IssueCode.get_label(code))
        cell_occ  = ws.cell(row=row, column=2, value=count)
        cell_pct  = ws.cell(row=row, column=3, value=f"{pct:.1f}%")

        _apply(cell_lib,  style_body)
        _apply(cell_occ,  style_pct)
        _apply(cell_pct,  style_pct)
        ws.row_dimensions[row].height = 18
        row += 1

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10


# ==============================================================================
# ONGLET DÉTAIL ANOMALIES
# ==============================================================================

def _write_anomalies_sheet(ws, df_export: pd.DataFrame) -> None:
    """
    Écrit l'onglet Détail anomalies.

    Colonnes : ID_SITE | NU_CNT | LIBELLE | DESCRIPTION
    - Retour à la ligne automatique sur LIBELLE et DESCRIPTION
    - Rangées alternées pour la lisibilité
    - Hauteur de ligne adaptée à la longueur du texte
    """
    ws.sheet_view.showGridLines = False

    if df_export.empty:
        ws.cell(row=1, column=1, value="Aucune anomalie détectée.")
        _apply(ws.cell(row=1, column=1), {"font": _font(size=11, bold=True), "align": _align()})
        return

    # En-tête
    labels = ["ID_SITE", "NU_CNT", "LIBELLE", "DESCRIPTION"]
    for col_idx, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        _apply(cell, _STYLE_HEADER)
    ws.row_dimensions[1].height = 24

    # Données
    for row_idx, row_data in enumerate(df_export.itertuples(index=False), 2):
        is_alt = (row_idx % 2 == 0)

        for col_idx, value in enumerate(row_data, 1):
            raw = _sanitize_excel_value(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=raw)

            if col_idx in (3, 4):  # LIBELLE et DESCRIPTION → wrap
                style = _STYLE_BODY_ALT if is_alt else _STYLE_BODY
            else:                   # ID_SITE, NU_CNT → centré, sans wrap
                style = {
                    "font":   _font(size=11),
                    "fill":   _fill(_COLOR_ROW_ALT) if is_alt else _fill("FFFFFF"),
                    "align":  _align(h="center", v="center"),
                    "border": _border_thin(),
                }
            _apply(cell, style)

        # Hauteur de ligne : estimation selon la longueur de la DESCRIPTION
        desc = str(row_data[3]) if row_data[3] else ""
        estimated_lines = max(1, len(desc) // 90 + 1)  # ~90 car / ligne selon la largeur
        ws.row_dimensions[row_idx].height = max(18, estimated_lines * 15)

    # Freeze première ligne
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D1"

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 14   # ID_SITE
    ws.column_dimensions["B"].width = 16   # NU_CNT
    ws.column_dimensions["C"].width = 42   # LIBELLE  (wrap)
    ws.column_dimensions["D"].width = 72   # DESCRIPTION (wrap, le plus large)


# ==============================================================================
# UTILITAIRE
# ==============================================================================

def _sanitize_excel_value(value):
    """Nettoie les caractères illégaux pour Excel."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value
