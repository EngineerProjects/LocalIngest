"""
Génération du rapport Excel.

Onglets :
- Résumé               : statistiques globales
- Sites avec anomalies : 1 ligne par site concerné
- Données enrichies    : 1 ligne par site (tous les sites actifs)
"""

from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill

from src.config import Config, ControlStatus, IssueCode, Priority, ValidationMode
from src.utils import format_number


_PRIORITY_SORT = {
    Priority.HIGH: 3,
    Priority.MEDIUM: 2,
    Priority.LOW: 1,
    Priority.NONE: 0,
}


def compute_global_stats(df: pd.DataFrame) -> dict:
    """Calcule les statistiques globales du rapport simplifié."""
    issue_counts = Counter()
    if "_ISSUE_CODES" in df.columns:
        for codes in df["_ISSUE_CODES"].fillna(""):
            for code in [
                item.strip() for item in str(codes).split(",") if item.strip()
            ]:
                issue_counts[code] += 1

    stats_by_country = {}
    if Config.COL_COUNTRY in df.columns:
        for country, grp in df.groupby(Config.COL_COUNTRY, dropna=False):
            stats_by_country[str(country)] = {
                "nb_sites": len(grp),
                "avec_anomalie": int((grp["_ISSUE_COUNT"] > 0).sum()),
                "gps_ok": int((grp["_CONTROL_MODE"] == ValidationMode.GPS).sum()),
                "adresse_ok": int(
                    (grp["_CONTROL_MODE"] == ValidationMode.ADDRESS).sum()
                ),
                "incomplets": int(
                    (grp["_CONTROL_MODE"] == ValidationMode.INCOMPLETE).sum()
                ),
            }

    return {
        "date_execution": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_sites": len(df),
        "sites_ok": int((df["_ISSUE_COUNT"] == 0).sum()),
        "sites_avec_anomalies": int((df["_ISSUE_COUNT"] > 0).sum()),
        "gps_ok": int((df["_CONTROL_MODE"] == ValidationMode.GPS).sum()),
        "adresse_ok": int((df["_CONTROL_MODE"] == ValidationMode.ADDRESS).sum()),
        "sites_incomplets": int(
            (df["_CONTROL_MODE"] == ValidationMode.INCOMPLETE).sum()
        ),
        "status_counts": {
            ControlStatus.OK: int((df["_CONTROL_STATUS"] == ControlStatus.OK).sum()),
            ControlStatus.TO_FIX: int(
                (df["_CONTROL_STATUS"] == ControlStatus.TO_FIX).sum()
            ),
            ControlStatus.TO_COMPLETE: int(
                (df["_CONTROL_STATUS"] == ControlStatus.TO_COMPLETE).sum()
            ),
            ControlStatus.TO_REVIEW: int(
                (df["_CONTROL_STATUS"] == ControlStatus.TO_REVIEW).sum()
            ),
        },
        "priority_counts": {
            Priority.HIGH: int((df["_PRIORITY"] == Priority.HIGH).sum()),
            Priority.MEDIUM: int((df["_PRIORITY"] == Priority.MEDIUM).sum()),
            Priority.LOW: int((df["_PRIORITY"] == Priority.LOW).sum()),
        },
        "issue_counts": dict(issue_counts.most_common()),
        "stats_by_country": stats_by_country,
    }


def generate_excel_report(
    df: pd.DataFrame,
    stats: dict,
    output_path: Path = None,
) -> Path:
    """Génère le rapport Excel simplifié (3 onglets, 1 ligne par site)."""
    print("\n" + "=" * 60)
    print("   GÉNÉRATION DU RAPPORT EXCEL")
    print("=" * 60)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Config.OUTPUT_DIR / f"rapport_anomalies_{timestamp}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5D50", end_color="2F5D50", fill_type="solid"
    )
    title_font = Font(bold=True, size=13)

    # Onglet 1 — Résumé
    _write_summary_sheet(wb.active, stats, header_font, header_fill, title_font)

    # Onglet 2 — Sites avec anomalies (1 ligne par site, uniquement anormaux)
    df_issues = _build_issues_export(df)
    _write_data_sheet(
        wb.create_sheet("Sites avec anomalies"),
        df_issues,
        header_font,
        header_fill,
    )

    # Onglet 3 — Données enrichies (1 ligne par site, tous)
    df_enriched = _build_enriched_export(df)
    _write_data_sheet(
        wb.create_sheet("Données enrichies"),
        df_enriched,
        header_font,
        header_fill,
    )

    wb.save(output_path)

    print(f"\nRapport généré : {output_path}")
    print(f"   Sites avec anomalies : {format_number(len(df_issues))} lignes")
    print(f"   Données enrichies    : {format_number(len(df_enriched))} lignes")

    return output_path


# ─────────────────────────────────────────────────────────────
# Constructeurs de vues
# ─────────────────────────────────────────────────────────────


def _build_issues_export(df: pd.DataFrame) -> pd.DataFrame:
    """Vue métier principale : 1 ligne par site en anomalie, triée par priorité."""
    export_df = df[df["_ISSUE_COUNT"] > 0].copy()
    export_df["_SORT_PRIORITY"] = export_df["_PRIORITY"].map(_PRIORITY_SORT).fillna(0)
    export_df = export_df.sort_values(
        by=[
            "_SORT_PRIORITY",
            "_ISSUE_COUNT",
            Config.COL_CONTRACT_ID,
            Config.COL_SITE_ID,
        ],
        ascending=[False, False, True, True],
    )
    return _build_site_anomalies_view(export_df)


def _build_site_anomalies_view(df: pd.DataFrame) -> pd.DataFrame:
    """Construit la vue avec les colonnes style 'Détail anomalies' mais 1 ligne par site."""
    rows = []

    for _, row in df.iterrows():
        codes = str(row.get("_ISSUE_CODES", "")).split(", ")
        codes = [c.strip() for c in codes if c.strip()]

        details = row.get("_ISSUE_DETAILS", [])
        if isinstance(details, list) and details:
            first_detail = details[0]

            consolidated_champs = "; ".join(
                set(
                    d.get("CHAMPS_CONCERNES", "")
                    for d in details
                    if d.get("CHAMPS_CONCERNES")
                )
            )
            consolidated_valeurs = "; ".join(
                set(
                    d.get("VALEUR_ACTUELLE", "")
                    for d in details
                    if d.get("VALEUR_ACTUELLE")
                )
            )
            consolidated_suggestions = "; ".join(
                set(d.get("SUGGESTION", "") for d in details if d.get("SUGGESTION"))
            )

            rows.append(
                {
                    "ID_SITE": row.get(Config.COL_SITE_ID),
                    "NU_CNT": row.get(Config.COL_CONTRACT_ID),
                    "CODE_ANOMALIE": ", ".join(codes),
                    "LIBELLE": row.get("_ISSUE_SUMMARY", ""),
                    "GRAVITE": row.get("_PRIORITY", ""),
                    "CATEGORIE": first_detail.get("CATEGORIE", ""),
                    "DESCRIPTION": "; ".join(
                        set(
                            d.get("DESCRIPTION", "")
                            for d in details
                            if d.get("DESCRIPTION")
                        )
                    ),
                    "CHAMPS_CONCERNES": consolidated_champs,
                    "VALEUR_ACTUELLE": consolidated_valeurs,
                    "SUGGESTION": consolidated_suggestions,
                    "SCORE_SIMILARITE": "",
                }
            )
        else:
            rows.append(
                {
                    "ID_SITE": row.get(Config.COL_SITE_ID),
                    "NU_CNT": row.get(Config.COL_CONTRACT_ID),
                    "CODE_ANOMALIE": ", ".join(codes),
                    "LIBELLE": row.get("_ISSUE_SUMMARY", ""),
                    "GRAVITE": row.get("_PRIORITY", ""),
                    "CATEGORIE": "",
                    "DESCRIPTION": "",
                    "CHAMPS_CONCERNES": "",
                    "VALEUR_ACTUELLE": "",
                    "SUGGESTION": "",
                    "SCORE_SIMILARITE": "",
                }
            )

    result_df = pd.DataFrame(rows)
    column_order = [
        "ID_SITE",
        "NU_CNT",
        "CODE_ANOMALIE",
        "LIBELLE",
        "GRAVITE",
        "CATEGORIE",
        "DESCRIPTION",
        "CHAMPS_CONCERNES",
        "VALEUR_ACTUELLE",
        "SUGGESTION",
        "SCORE_SIMILARITE",
    ]
    return (
        result_df[column_order]
        if all(c in result_df.columns for c in column_order)
        else result_df
    )


def _build_enriched_export(df: pd.DataFrame) -> pd.DataFrame:
    """Vue enrichie complète : 1 ligne par site actif."""
    export_df = df.copy()
    export_df["_SORT_PRIORITY"] = export_df["_PRIORITY"].map(_PRIORITY_SORT).fillna(0)
    export_df = export_df.sort_values(
        by=[
            "_SORT_PRIORITY",
            "_ISSUE_COUNT",
            Config.COL_CONTRACT_ID,
            Config.COL_SITE_ID,
        ],
        ascending=[False, False, True, True],
    )
    columns = [
        col for col in Config.ENRICHED_EXPORT_COLUMNS if col in export_df.columns
    ]
    return export_df[columns].where(pd.notna(export_df[columns]), None)


# ─────────────────────────────────────────────────────────────
# Écriture des onglets
# ─────────────────────────────────────────────────────────────


def _write_summary_sheet(ws, stats: dict, header_font, header_fill, title_font) -> None:
    """Écrit l'onglet résumé."""
    ws.title = "Résumé"
    row = 1

    def write_title(text: str) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = title_font
        row += 2

    def write_kv(label: str, value) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    def write_header(values) -> None:
        nonlocal row
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

    write_title("SECTION A — Vue globale")
    for label, value in [
        ("Date d'exécution", stats["date_execution"]),
        ("Fichier source", Config.INPUT_FILE.name),
        ("", ""),
        ("Sites actifs analysés", format_number(stats["total_sites"])),
        ("", ""),
        ("Sites sans anomalie (Total validés)", format_number(stats["sites_ok"])),
        ("  - validés par GPS", format_number(stats["gps_ok"])),
        ("  - validés par adresse", format_number(stats["adresse_ok"])),
        ("", ""),
        ("Sites avec anomalie(s)", format_number(stats["sites_avec_anomalies"])),
    ]:
        write_kv(label, value)

    row += 1
    write_title("SECTION B — Répartition par type d'anomalie")
    write_header(["Code", "Libellé", "Occurrences", "Priorité"])
    for code, count in stats["issue_counts"].items():
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=IssueCode.get_label(code))
        ws.cell(row=row, column=3, value=count)
        ws.cell(row=row, column=4, value=IssueCode.get_priority(code))
        row += 1

    row += 1
    write_title("SECTION C — Répartition par pays")
    write_header(
        ["Pays", "Nb sites", "Avec anomalies", "GPS OK", "Adresse OK", "Incomplets"]
    )
    for country, values in sorted(stats["stats_by_country"].items()):
        ws.cell(row=row, column=1, value=country)
        ws.cell(row=row, column=2, value=values["nb_sites"])
        ws.cell(row=row, column=3, value=values["avec_anomalie"])
        ws.cell(row=row, column=4, value=values["gps_ok"])
        ws.cell(row=row, column=5, value=values["adresse_ok"])
        ws.cell(row=row, column=6, value=values["incomplets"])
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


def _write_data_sheet(ws, df_export: pd.DataFrame, header_font, header_fill) -> None:
    """Écrit un onglet tabulaire standard (1 ligne par site)."""
    if df_export.empty:
        ws.cell(row=1, column=1, value="Aucune donnée à afficher.")
        return

    for col_idx, col_name in enumerate(df_export.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row_data in enumerate(df_export.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=_sanitize_excel_value(value))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    width_map = {
        Config.COL_CONTRACT_ID: 16,
        Config.COL_SITE_ID: 12,
        Config.COL_CLIENT_NAME: 26,
        Config.COL_SITE_NAME: 26,
        Config.COL_COUNTRY: 16,
        Config.COL_POSTAL_CODE: 12,
        Config.COL_CITY: 22,
        Config.COL_STREET_NUMBER: 14,
        Config.COL_STREET_NAME: 28,
        Config.COL_STREET_FULL: 28,
        Config.COL_FULL_ADDRESS: 28,
        Config.COL_LONGITUDE: 14,
        Config.COL_LATITUDE: 14,
        "_CONTROL_MODE": 16,
        "_CONTROL_STATUS": 16,
        "_PRIORITY": 12,
        "_ISSUE_COUNT": 12,
        "_ISSUE_CODES": 22,
        "_ISSUE_SUMMARY": 52,
        "_SITE_KEY": 24,
        "_HAS_GPS": 12,
        "_GPS_IN_COUNTRY": 16,
        "_GPS_MATCH_POSTAL_CODE": 22,
        "_ADDRESS_PRESENT": 16,
        "_STREET_NUMBER_PRESENT": 22,
        "_DEPARTMENT_CODE": 18,
        "_DEPARTMENT_COUNTRY_OK": 24,
    }

    for col_idx, col_name in enumerate(df_export.columns, 1):
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = width_map.get(col_name, 18)


def _sanitize_excel_value(value):
    """Nettoie les caractères illégaux avant écriture dans Excel."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value
