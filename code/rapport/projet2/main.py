"""
Outil de contrôle qualité des données Proxima
----------------------------------------------
Usage : python main.py --path "/chemin/vers/fichier.csv"

Pipeline :
    1. Chargement et filtrage des contrats actifs
    2. Validation des adresses via API Adresse (data.gouv.fr)
    3. Vérification de la cohérence des coordonnées GPS
    4. Validation des données techniques (capitaux / superficie)
    5. Génération du rapport Excel des anomalies
"""

import argparse
import sys
import time
import requests
import pandas as pd
from geopy.distance import geodesic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path


# =============================================================================
# CONFIGURATION GÉNÉRALE
# =============================================================================

# Colonnes attendues dans le fichier CSV source (à adapter si nécessaire)
COL_ID             = "id_point_gestion"
COL_STATUT         = "cdsitp"
COL_DATE_RESIL     = "dtresilp"
COL_ADRESSE        = "adresse"
COL_CODE_POSTAL    = "code_postal"
COL_VILLE          = "ville"
COL_PAYS           = "pays"
COL_LATITUDE       = "latitude"
COL_LONGITUDE      = "longitude"
COL_CAPITAUX       = "capitaux"
COL_SUPERFICIE     = "superficie"

# Seuil de score de confiance de l'API Adresse (entre 0 et 1)
# En dessous de ce seuil, l'adresse est considérée comme non reconnue
SCORE_GEOCODAGE_MIN = 0.6

# Distance maximale tolérée (en km) entre le GPS du fichier et le GPS retourné par l'API
DISTANCE_GPS_MAX_KM = 50

# Plages de valeurs GPS valides pour la France métropolitaine
LAT_MIN, LAT_MAX   = 41.0, 51.5
LON_MIN, LON_MAX   = -5.5, 10.0

# Délai entre deux appels à l'API (en secondes) pour éviter de saturer le service
API_DELAI_SECONDES = 0.1

# URL de l'API Adresse du gouvernement
API_ADRESSE_URL = "https://api-adresse.data.gouv.fr/search/"


# =============================================================================
# ÉTAPE 1 — CHARGEMENT ET FILTRAGE DES DONNÉES
# =============================================================================

def charger_et_filtrer(chemin_csv: str) -> pd.DataFrame:
    """
    Charge le fichier CSV et conserve uniquement les contrats actifs :
        - cdsitp == 1
        - dtresilp vide (pas de date de résiliation)

    Returns:
        DataFrame filtré avec uniquement les contrats actifs.
    """
    print(f"\n[Étape 1] Chargement du fichier : {chemin_csv}")

    try:
        df = pd.read_csv(chemin_csv, sep=None, engine="python", dtype=str)
    except FileNotFoundError:
        print(f"  ERREUR : Fichier introuvable → {chemin_csv}")
        sys.exit(1)
    except Exception as e:
        print(f"  ERREUR lors de la lecture du fichier : {e}")
        sys.exit(1)

    print(f"  Lignes chargées          : {len(df)}")

    # Vérification de la présence des colonnes minimales requises
    colonnes_requises = [
        COL_ID, COL_STATUT, COL_DATE_RESIL,
        COL_ADRESSE, COL_CODE_POSTAL, COL_VILLE,
        COL_LATITUDE, COL_LONGITUDE
    ]
    colonnes_manquantes = [c for c in colonnes_requises if c not in df.columns]
    if colonnes_manquantes:
        print(f"  ERREUR : Colonnes manquantes dans le CSV → {colonnes_manquantes}")
        sys.exit(1)

    # Nettoyage des espaces superflus sur les colonnes texte
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    # Filtre 1 : cdsitp == 1
    df[COL_STATUT] = df[COL_STATUT].astype(str)
    df_actifs = df[df[COL_STATUT] == "1"].copy()

    # Filtre 2 : dtresilp vide
    df_actifs = df_actifs[
        df_actifs[COL_DATE_RESIL].isna() | (df_actifs[COL_DATE_RESIL] == "")
    ].copy()

    print(f"  Contrats actifs retenus  : {len(df_actifs)}")
    print(f"  Contrats exclus (résiliés ou inactifs) : {len(df) - len(df_actifs)}")

    return df_actifs.reset_index(drop=True)


# =============================================================================
# ÉTAPE 2 — VALIDATION DES ADRESSES VIA GÉOCODAGE
# =============================================================================

def geocoder_adresse(adresse: str, code_postal: str, ville: str) -> dict:
    """
    Appelle l'API Adresse du gouvernement pour géocoder une adresse.

    Returns:
        dict avec les clés :
            - score       : score de confiance (float entre 0 et 1), ou None si échec
            - lat_api     : latitude retournée par l'API
            - lon_api     : longitude retournée par l'API
            - label       : adresse normalisée retournée par l'API
    """
    texte = f"{adresse} {code_postal} {ville}".strip()
    params = {"q": texte, "limit": 1}

    try:
        reponse = requests.get(API_ADRESSE_URL, params=params, timeout=5)
        reponse.raise_for_status()
        donnees = reponse.json()

        if donnees.get("features"):
            feature  = donnees["features"][0]
            props    = feature["properties"]
            geometry = feature["geometry"]
            return {
                "score"   : props.get("score"),
                "lat_api" : geometry["coordinates"][1],
                "lon_api" : geometry["coordinates"][0],
                "label"   : props.get("label", "")
            }
    except requests.exceptions.Timeout:
        pass
    except requests.exceptions.RequestException:
        pass

    return {"score": None, "lat_api": None, "lon_api": None, "label": ""}


def valider_adresses(df: pd.DataFrame) -> list:
    """
    Géocode toutes les adresses du DataFrame et retourne la liste des anomalies détectées.

    Returns:
        Liste de dicts décrivant chaque anomalie, et DataFrame enrichi avec les
        coordonnées GPS de référence issues de l'API.
    """
    print(f"\n[Étape 2] Validation des adresses via API Adresse (data.gouv.fr)")
    print(f"  Adresses à traiter : {len(df)}")

    anomalies   = []
    lat_api_col = []
    lon_api_col = []

    for i, row in df.iterrows():
        adresse     = str(row.get(COL_ADRESSE, ""))
        code_postal = str(row.get(COL_CODE_POSTAL, ""))
        ville       = str(row.get(COL_VILLE, ""))
        id_pg       = row.get(COL_ID, f"ligne_{i}")

        # Adresse complètement vide → anomalie directe sans appel API
        if not adresse or adresse.lower() in ("nan", ""):
            anomalies.append({
                "id_point_gestion" : id_pg,
                "type_anomalie"    : "Adresse vide",
                "valeur_constatee" : "Champ adresse absent ou vide",
                "detail"           : ""
            })
            lat_api_col.append(None)
            lon_api_col.append(None)
            continue

        resultat = geocoder_adresse(adresse, code_postal, ville)
        lat_api_col.append(resultat["lat_api"])
        lon_api_col.append(resultat["lon_api"])

        # Échec total de l'API
        if resultat["score"] is None:
            anomalies.append({
                "id_point_gestion" : id_pg,
                "type_anomalie"    : "Adresse non reconnue (erreur API)",
                "valeur_constatee" : f"{adresse}, {code_postal} {ville}",
                "detail"           : "L'API n'a pas retourné de résultat"
            })

        # Score insuffisant → adresse douteuse (faute de frappe, ville inconnue, etc.)
        elif resultat["score"] < SCORE_GEOCODAGE_MIN:
            anomalies.append({
                "id_point_gestion" : id_pg,
                "type_anomalie"    : "Adresse non reconnue (score insuffisant)",
                "valeur_constatee" : f"{adresse}, {code_postal} {ville}",
                "detail"           : f"Score obtenu : {resultat['score']:.2f} "
                                     f"(seuil minimum : {SCORE_GEOCODAGE_MIN})"
            })

        time.sleep(API_DELAI_SECONDES)

        if (i + 1) % 50 == 0:
            print(f"  ... {i + 1} / {len(df)} adresses traitées")

    df["_lat_api"] = lat_api_col
    df["_lon_api"] = lon_api_col

    print(f"  Anomalies d'adresse détectées : {len(anomalies)}")
    return anomalies, df


# =============================================================================
# ÉTAPE 3 — VÉRIFICATION DE LA COHÉRENCE DES COORDONNÉES GPS
# =============================================================================

def valider_gps(df: pd.DataFrame) -> list:
    """
    Vérifie la cohérence des coordonnées GPS présentes dans le fichier :
        - Valeurs dans les plages réalistes pour la France métropolitaine
        - Écart entre le GPS du fichier et le GPS retourné par l'API

    Returns:
        Liste de dicts décrivant chaque anomalie GPS détectée.
    """
    print(f"\n[Étape 3] Vérification de la cohérence des coordonnées GPS")

    anomalies = []

    for i, row in df.iterrows():
        id_pg = row.get(COL_ID, f"ligne_{i}")
        lat_str = row.get(COL_LATITUDE, "")
        lon_str = row.get(COL_LONGITUDE, "")

        # Conversion des coordonnées GPS du fichier
        try:
            lat_fichier = float(str(lat_str).replace(",", "."))
            lon_fichier = float(str(lon_str).replace(",", "."))
        except (ValueError, TypeError):
            anomalies.append({
                "id_point_gestion" : id_pg,
                "type_anomalie"    : "Coordonnées GPS invalides (format)",
                "valeur_constatee" : f"lat={lat_str}, lon={lon_str}",
                "detail"           : "Impossible de convertir les valeurs en nombre"
            })
            continue

        # Vérification des plages réalistes (France métropolitaine)
        hors_plage = (
            not (LAT_MIN <= lat_fichier <= LAT_MAX) or
            not (LON_MIN <= lon_fichier <= LON_MAX)
        )
        if hors_plage:
            anomalies.append({
                "id_point_gestion" : id_pg,
                "type_anomalie"    : "Coordonnées GPS hors plage France",
                "valeur_constatee" : f"lat={lat_fichier}, lon={lon_fichier}",
                "detail"           : f"Plages attendues : lat [{LAT_MIN}, {LAT_MAX}] "
                                     f"/ lon [{LON_MIN}, {LON_MAX}]"
            })
            continue

        # Comparaison avec les coordonnées retournées par l'API (étape 2)
        lat_api = row.get("_lat_api")
        lon_api = row.get("_lon_api")

        if lat_api is not None and lon_api is not None:
            try:
                distance_km = geodesic(
                    (lat_fichier, lon_fichier),
                    (lat_api, lon_api)
                ).kilometers

                if distance_km > DISTANCE_GPS_MAX_KM:
                    anomalies.append({
                        "id_point_gestion" : id_pg,
                        "type_anomalie"    : "Incohérence géographique (écart GPS)",
                        "valeur_constatee" : f"lat={lat_fichier}, lon={lon_fichier}",
                        "detail"           : f"Écart de {distance_km:.1f} km avec "
                                             f"la position attendue (max : {DISTANCE_GPS_MAX_KM} km)"
                    })
            except Exception:
                pass

    print(f"  Anomalies GPS détectées : {len(anomalies)}")
    return anomalies


# =============================================================================
# ÉTAPE 4 — VALIDATION DES DONNÉES TECHNIQUES
# =============================================================================

def valider_donnees_techniques(df: pd.DataFrame) -> list:
    """
    Vérifie que chaque contrat possède au moins un champ parmi :
        - capitaux
        - superficie

    Returns:
        Liste de dicts décrivant chaque anomalie technique détectée.
    """
    print(f"\n[Étape 4] Validation des données techniques (capitaux / superficie)")

    anomalies = []

    col_capitaux   = COL_CAPITAUX   in df.columns
    col_superficie = COL_SUPERFICIE in df.columns

    if not col_capitaux and not col_superficie:
        print("  AVERTISSEMENT : colonnes 'capitaux' et 'superficie' absentes du fichier, "
              "étape ignorée.")
        return anomalies

    for i, row in df.iterrows():
        id_pg = row.get(COL_ID, f"ligne_{i}")

        capitaux_vide   = True
        superficie_vide = True

        if col_capitaux:
            val = row.get(COL_CAPITAUX, "")
            capitaux_vide = pd.isna(val) or str(val).strip() in ("", "nan", "0")

        if col_superficie:
            val = row.get(COL_SUPERFICIE, "")
            superficie_vide = pd.isna(val) or str(val).strip() in ("", "nan", "0")

        if capitaux_vide and superficie_vide:
            anomalies.append({
                "id_point_gestion" : id_pg,
                "type_anomalie"    : "Données techniques manquantes",
                "valeur_constatee" : "capitaux=vide, superficie=vide",
                "detail"           : "Au moins l'un des deux champs doit être renseigné"
            })

    print(f"  Anomalies techniques détectées : {len(anomalies)}")
    return anomalies


# =============================================================================
# ÉTAPE 5 — GÉNÉRATION DU RAPPORT EXCEL
# =============================================================================

def generer_rapport_excel(toutes_anomalies: list, dossier_sortie: str) -> str:
    """
    Génère un fichier Excel horodaté recensant toutes les anomalies détectées,
    classées par point de gestion.

    Returns:
        Chemin du fichier Excel généré.
    """
    print(f"\n[Étape 5] Génération du rapport Excel")

    if not toutes_anomalies:
        print("  Aucune anomalie détectée. Aucun rapport généré.")
        return None

    # Tri par point de gestion puis par type d'anomalie
    df_anomalies = pd.DataFrame(toutes_anomalies).sort_values(
        by=["id_point_gestion", "type_anomalie"]
    )

    # Nom du fichier horodaté
    timestamp   = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nom_fichier = f"rapport_anomalies_{timestamp}.xlsx"
    chemin      = Path(dossier_sortie) / nom_fichier

    # Construction du fichier Excel avec mise en forme
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"

    # --- En-têtes ---
    entetes = [
        "Point de gestion",
        "Type d'anomalie",
        "Valeur constatée",
        "Détail",
        "Date d'analyse"
    ]
    couleur_entete = "2F4F8F"
    style_entete   = Font(bold=True, color="FFFFFF")

    for col_idx, titre in enumerate(entetes, start=1):
        cellule = ws.cell(row=1, column=col_idx, value=titre)
        cellule.font      = style_entete
        cellule.fill      = PatternFill("solid", fgColor=couleur_entete)
        cellule.alignment = Alignment(horizontal="center", vertical="center")

    # Alternance de couleurs pour les lignes
    couleur_paire   = "DCE6F1"
    couleur_impaire = "FFFFFF"
    date_analyse    = datetime.now().strftime("%d/%m/%Y")

    for row_idx, (_, row) in enumerate(df_anomalies.iterrows(), start=2):
        couleur = couleur_paire if row_idx % 2 == 0 else couleur_impaire
        remplissage = PatternFill("solid", fgColor=couleur)

        valeurs = [
            row.get("id_point_gestion", ""),
            row.get("type_anomalie", ""),
            row.get("valeur_constatee", ""),
            row.get("detail", ""),
            date_analyse
        ]
        for col_idx, valeur in enumerate(valeurs, start=1):
            cellule      = ws.cell(row=row_idx, column=col_idx, value=valeur)
            cellule.fill = remplissage

    # Ajustement automatique de la largeur des colonnes
    largeurs = [20, 40, 40, 60, 18]
    for col_idx, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = largeur

    # Onglet de synthèse
    ws_synthese = wb.create_sheet(title="Synthèse")
    types_anomalies = df_anomalies["type_anomalie"].value_counts().reset_index()
    types_anomalies.columns = ["Type d'anomalie", "Nombre"]

    ws_synthese.cell(row=1, column=1, value="Type d'anomalie").font = Font(bold=True)
    ws_synthese.cell(row=1, column=2, value="Nombre").font           = Font(bold=True)

    for r, row in enumerate(types_anomalies.itertuples(index=False), start=2):
        ws_synthese.cell(row=r, column=1, value=row[0])
        ws_synthese.cell(row=r, column=2, value=row[1])

    ws_synthese.cell(
        row=len(types_anomalies) + 3,
        column=1,
        value=f"Total anomalies : {len(df_anomalies)}"
    ).font = Font(bold=True)

    ws_synthese.column_dimensions["A"].width = 45
    ws_synthese.column_dimensions["B"].width = 12

    wb.save(chemin)
    print(f"  Rapport généré : {chemin}")
    print(f"  Total anomalies : {len(df_anomalies)}")

    return str(chemin)


# =============================================================================
# POINT D'ENTRÉE
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Contrôle qualité des données Proxima"
    )
    parser.add_argument(
        "--path",
        required=True,
        help="Chemin vers le fichier CSV à analyser"
    )
    parser.add_argument(
        "--output",
        default=".",
        help="Dossier de sortie pour le rapport Excel (défaut : répertoire courant)"
    )
    args = parser.parse_args()

    print("=" * 60)
    print("  Contrôle qualité des données Proxima")
    print(f"  Démarré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    print("=" * 60)

    # Étape 1 : Chargement et filtrage
    df_actifs = charger_et_filtrer(args.path)

    # Étape 2 : Validation des adresses
    anomalies_adresses, df_enrichi = valider_adresses(df_actifs)

    # Étape 3 : Vérification GPS
    anomalies_gps = valider_gps(df_enrichi)

    # Étape 4 : Validation des données techniques
    anomalies_techniques = valider_donnees_techniques(df_enrichi)

    # Consolidation de toutes les anomalies
    toutes_anomalies = anomalies_adresses + anomalies_gps + anomalies_techniques

    # Étape 5 : Génération du rapport Excel
    generer_rapport_excel(toutes_anomalies, args.output)

    print("\n" + "=" * 60)
    print("  Pipeline terminé avec succès.")
    print(f"  Récapitulatif :")
    print(f"    - Anomalies adresses   : {len(anomalies_adresses)}")
    print(f"    - Anomalies GPS        : {len(anomalies_gps)}")
    print(f"    - Anomalies techniques : {len(anomalies_techniques)}")
    print(f"    - Total                : {len(toutes_anomalies)}")
    print("=" * 60)


if __name__ == "__main__":
    main()